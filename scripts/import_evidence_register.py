"""
Import material evidence + prices from:
  docs/Materials/material_evidence_register_lv.xlsx

Run from repo root:
  python scripts/import_evidence_register.py
"""

import sys
import os
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from datetime import date
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs" / "Materials" / "material_evidence_register_lv.xlsx"

sys.path.insert(0, str(ROOT / "backend"))
os.environ.setdefault("DATABASE_URL", "postgresql://pod:pod_secret@localhost:5432/pod_mfg")

from app.models import MaterialLibrary, MaterialPrice  # noqa: E402

DATABASE_URL = os.environ["DATABASE_URL"]
engine = create_engine(DATABASE_URL, pool_pre_ping=True)

# ── Evidence status rule ──────────────────────────────────────────────────────

def compute_status(manufacturer, supplier_url, datasheet_url, lambda_val):
    has_ds  = bool(datasheet_url)
    has_sup = bool(supplier_url)
    has_mfr = bool(manufacturer)
    has_lam = lambda_val is not None and float(lambda_val) > 0
    if has_ds and has_mfr and has_lam and has_sup:
        return "verified"
    elif has_ds or has_sup:
        return "partial"
    return "missing"


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (date,)):
        return val
    try:
        return date.fromisoformat(str(val))
    except Exception:
        return None


def clean(val):
    """Return None for empty/whitespace strings."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ── Load workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(str(XLSX), data_only=True)

# ── Sheet 1: Evidence Register ────────────────────────────────────────────────

ev_ws   = wb["Evidence Register"]
ev_rows = list(ev_ws.iter_rows(values_only=True))
ev_hdr  = [str(h).strip() if h else "" for h in ev_rows[0]]

def evcol(row, name):
    try:
        return row[ev_hdr.index(name)]
    except (ValueError, IndexError):
        return None

evidence_data = {}   # supplier_ref → dict of fields
for row in ev_rows[1:]:
    ref = clean(evcol(row, "App supplier_ref"))
    if not ref:
        continue
    evidence_data[ref] = {
        "manufacturer":     clean(evcol(row, "Manufacturer")),
        "supplier_name":    clean(evcol(row, "Supplier / seller")),
        "supplier_url":     clean(evcol(row, "Supplier URL")),
        "datasheet_url":    clean(evcol(row, "Data sheet URL")),
        "dop_url":          clean(evcol(row, "DoP / certificate URL")),
        "lambda_W_mK":      evcol(row, "λ W/mK"),
        "fire_euroclass":   clean(evcol(row, "Fire / Euroclass")),
        "density_kg_m3":    evcol(row, "Density kg/m³"),
        "price_source_url": clean(evcol(row, "Price source URL")),
        "price_checked_at": parse_date(evcol(row, "Checked date")),
        "notes":            clean(evcol(row, "Notes")),
    }

# ── Sheet 2: Price Inputs ─────────────────────────────────────────────────────

pr_ws   = wb["Price Inputs"]
pr_rows = list(pr_ws.iter_rows(values_only=True))
pr_hdr  = [str(h).strip() if h else "" for h in pr_rows[0]]

def prcol(row, name):
    try:
        return row[pr_hdr.index(name)]
    except (ValueError, IndexError):
        return None

price_data = {}   # supplier_ref → list of price dicts
for row in pr_rows[1:]:
    ref = clean(prcol(row, "supplier_ref"))
    if not ref:
        continue
    unit       = clean(prcol(row, "default_unit"))
    unit_price = prcol(row, "unit_price")
    currency   = clean(prcol(row, "currency")) or "EUR"
    price_type = clean(prcol(row, "price_type")) or "retail_lv"
    source_url = clean(prcol(row, "source_url"))
    checked    = parse_date(prcol(row, "checked_date"))
    notes      = clean(prcol(row, "notes"))

    # Handle combined unit like "m2/lm" — split into two records
    units = [u.strip() for u in (unit or "").split("/") if u.strip()]

    for u in units:
        price_data.setdefault(ref, []).append({
            "unit":             u,
            "price_per_unit":   float(unit_price) if unit_price is not None else None,
            "currency":         currency,
            "price_type":       price_type,
            "price_source_url": source_url,
            "price_checked_at": checked,
            "notes":            notes,
        })

# ── Apply to DB ───────────────────────────────────────────────────────────────

with Session(engine) as db:
    materials = db.query(MaterialLibrary).all()
    mat_by_ref = {m.supplier_ref: m for m in materials if m.supplier_ref}

    ev_updated = 0
    ev_skipped = 0
    pr_updated = 0
    pr_skipped = 0

    for ref, ev in evidence_data.items():
        mat = mat_by_ref.get(ref)
        if mat is None:
            print(f"  [SKIP evidence] no material with supplier_ref={ref!r}")
            ev_skipped += 1
            continue

        # Update evidence fields
        if ev["manufacturer"]     : mat.manufacturer     = ev["manufacturer"]
        if ev["supplier_name"]    : mat.supplier_name    = ev["supplier_name"]
        if ev["supplier_url"]     : mat.supplier_url     = ev["supplier_url"]
        if ev["datasheet_url"]    : mat.datasheet_url    = ev["datasheet_url"]
        if ev["dop_url"]          : mat.dop_url          = ev["dop_url"]
        if ev["fire_euroclass"]   : mat.fire_euroclass   = ev["fire_euroclass"]
        if ev["density_kg_m3"] is not None: mat.density_kg_m3 = float(ev["density_kg_m3"])
        if ev["price_source_url"] : mat.price_source_url = ev["price_source_url"]
        if ev["price_checked_at"] : mat.price_checked_at = ev["price_checked_at"]
        if ev["lambda_W_mK"] is not None:
            mat.lambda_W_mK = float(ev["lambda_W_mK"])

        # Recompute evidence_status
        mat.evidence_status = compute_status(
            mat.manufacturer, mat.supplier_url, mat.datasheet_url, mat.lambda_W_mK
        )

        ev_updated += 1
        print(f"  [evidence] {mat.name!r:<50} → {mat.evidence_status}")

    db.flush()

    # Prices — replace all prices for each ref then insert from register
    for ref, price_rows in price_data.items():
        mat = mat_by_ref.get(ref)
        if mat is None:
            print(f"  [SKIP price] no material with supplier_ref={ref!r}")
            pr_skipped += 1
            continue

        for pr in price_rows:
            if pr["price_per_unit"] is None:
                continue

            # Check if price record already exists for this unit
            existing = next(
                (p for p in mat.prices if p.unit == pr["unit"]),
                None
            )
            if existing:
                existing.price_per_unit  = pr["price_per_unit"]
                existing.currency        = pr["currency"]
                existing.price_type      = pr["price_type"]
                existing.notes           = pr["notes"]
                existing.is_default      = True
                action = "update"
            else:
                new_price = MaterialPrice(
                    material_id    = mat.id,
                    unit           = pr["unit"],
                    price_per_unit = pr["price_per_unit"],
                    currency       = pr["currency"],
                    price_type     = pr["price_type"],
                    notes          = pr["notes"],
                    is_default     = True,
                )
                db.add(new_price)
                action = "insert"

            pr_updated += 1
            print(f"  [price {action}] {mat.name!r:<50} {pr['unit']:>4} @ €{pr['price_per_unit']}")

    db.commit()
    print(f"\n✓ Evidence: {ev_updated} updated, {ev_skipped} skipped (no DB match)")
    print(f"✓ Prices:   {pr_updated} updated/inserted, {pr_skipped} skipped")
