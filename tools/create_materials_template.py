"""
Generate tools/materials_template.xlsx with a starter set of Nordic/Latvian materials.

Run once:  python tools/create_materials_template.py
Then fill in your own rows and import with:
    python tools/import_materials.py --file tools/materials_template.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(__file__), "materials_template.xlsx")

# ── colour palette ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
REQ_FILL    = PatternFill("solid", fgColor="2C5F8A")   # required columns
OPT_FILL    = PatternFill("solid", fgColor="3D7AB5")   # optional columns
BOLD_WHITE  = Font(bold=True, color="FFFFFF")
CENTRE      = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Materials sheet ──────────────────────────────────────────────────────────
MAT_COLS = [
    # (header, required, width)
    ("name",                          True,  32),
    ("manufacturer",                  False, 22),
    ("spec_ref",                      False, 18),
    ("lambda_W_mK",                   True,  14),   # thermal conductivity
    ("density_kg_m3",                 False, 14),
    ("cp_J_kgK",                      False, 12),   # specific heat capacity
    ("fire_euroclass",                False, 14),   # A1, A2, B, C, D, E, F
    ("embodied_carbon_kgCO2e_per_kg", False, 24),
    ("price_per_unit",                False, 14),
    ("unit",                          False, 10),   # m2, m3, kg, m, pcs
    ("currency",                      False, 10),
    ("supplier_ref",                  False, 18),
]

MATERIALS = [
    # Structural timber ────────────────────────────────────────────────────────
    ("KVH C24 47×147 (Latvian)",     "Latvijas Finieris / Binderholz", "EN 14081",  0.13, 500,  1600, "D",  0.263, 0.85, "m",  "EUR", "LV-KVH-C24-147"),
    ("KVH C24 47×195 (Latvian)",     "Latvijas Finieris / Binderholz", "EN 14081",  0.13, 500,  1600, "D",  0.263, 1.10, "m",  "EUR", "LV-KVH-C24-195"),
    ("KVH C24 47×220 (Latvian)",     "Latvijas Finieris / Binderholz", "EN 14081",  0.13, 500,  1600, "D",  0.263, 1.25, "m",  "EUR", "LV-KVH-C24-220"),
    ("KVH C24 47×97 Noggin (Latvian)","Latvijas Finieris / Binderholz","EN 14081",  0.13, 500,  1600, "D",  0.263, 0.55, "m",  "EUR", "LV-KVH-C24-97"),

    # Structural boards ────────────────────────────────────────────────────────
    ("OSB/3 12mm (Egger)",           "Egger",                          "EN 300",    0.13, 600,  1700, "D",  0.42,  8.50, "m2", "EUR", "EGGER-OSB3-12"),
    ("OSB/3 18mm floor (Egger)",     "Egger",                          "EN 300",    0.13, 600,  1700, "D",  0.42, 12.00, "m2", "EUR", "EGGER-OSB3-18"),

    # Insulation — mineral wool ─────────────────────────────────────────────────
    ("Rockwool Flexi 45 (between-stud)", "Rockwool SE",               "EN 13162",  0.037, 45,  840,  "A1", 1.28,  6.80, "m2", "EUR", "RW-FLEXI-45"),
    ("Rockwool Flexi 100",           "Rockwool SE",                    "EN 13162",  0.037, 45,  840,  "A1", 1.28,  9.40, "m2", "EUR", "RW-FLEXI-100"),
    ("Rockwool Frontrock Max E 100", "Rockwool SE",                    "EN 13162",  0.036, 160, 840,  "A2", 1.35, 18.50, "m2", "EUR", "RW-FRONTROCK-100"),
    ("Paroc eXtra 100 (between-rafter)", "Paroc SE",                  "EN 13162",  0.033, 30,  840,  "A1", 1.22,  8.20, "m2", "EUR", "PAROC-EXTRA-100"),
    ("ISOVER Multimax 30 (roof)",    "Saint-Gobain ISOVER SE",         "EN 13162",  0.033, 30,  840,  "A1", 1.22,  7.90, "m2", "EUR", "ISOVER-MM30"),

    # Insulation — PIR/PUR ────────────────────────────────────────────────────
    ("Kingspan Kooltherm K15 50mm",  "Kingspan Nordic",                "EN 13166",  0.020, 30,  1450, "F",  2.85, 22.00, "m2", "EUR", "KS-K15-50"),
    ("Kingspan Kooltherm K15 80mm",  "Kingspan Nordic",                "EN 13166",  0.020, 30,  1450, "F",  2.85, 30.00, "m2", "EUR", "KS-K15-80"),
    ("Kingspan Kooltherm K15 100mm", "Kingspan Nordic",                "EN 13166",  0.020, 30,  1450, "F",  2.85, 36.00, "m2", "EUR", "KS-K15-100"),

    # Plasterboard ──────────────────────────────────────────────────────────────
    ("Gyproc Standard 12.5mm",       "Saint-Gobain Gyproc SE",         "EN 520",    0.25,  825, 1000, "A2", 0.39,  4.50, "m2", "EUR", "GYPROC-STD-125"),
    ("Gyproc FireLine 15mm",         "Saint-Gobain Gyproc SE",         "EN 520",    0.25,  825, 1000, "A2", 0.39,  5.80, "m2", "EUR", "GYPROC-FL-15"),

    # Membranes ────────────────────────────────────────────────────────────────
    ("Siga Majrex 200 VCL",          "Siga",                           "EN 13984",  None,  None, None, "B",  2.10,  1.85, "m2", "EUR", "SIGA-MAJREX-200"),
    ("Tyvek Housewrap Breather",     "DuPont Tyvek",                   "EN 13859-2",None,  None, None, "B",  2.80,  1.40, "m2", "EUR", "TYVEK-HW"),

    # Acoustic ──────────────────────────────────────────────────────────────────
    ("Rockwool Acoustic 45 (party wall)","Rockwool SE",               "EN 13162",  0.034, 45,  840,  "A1", 1.28,  7.20, "m2", "EUR", "RW-ACOUSTIC-45"),

    # External cladding ────────────────────────────────────────────────────────
    ("Latvian Spruce Feather-edge 21mm","Latvijas Finieris",           "EN 14915",  0.13,  450, 1600, "D",  0.41,  9.50, "m2", "EUR", "LV-SPRUCE-FE-21"),
    ("Cembrit Patina Fibre-cement",  "Cembrit Nordic",                 "EN 12467",  0.58, 1400, 840,  "A2", 0.89, 32.00, "m2", "EUR", "CEMBRIT-PATINA"),
]

# ── Junctions sheet ──────────────────────────────────────────────────────────
JNC_COLS = [
    ("code",                      True,  18),
    ("type",                      True,  18),
    ("psi_value_W_mK",            True,  16),
    ("psi_source",                True,  18),
    ("build_up_type",             False, 18),
    ("insulation_continuity",     False, 20),
    ("thermal_break_present",     False, 20),
    ("min_outboard_insulation_mm",False, 24),
    ("passivhaus_flag",           False, 16),
    ("cert_ref",                  False, 24),
]

JUNCTIONS = [
    # code              type            psi     source          build_up        cont    tb      min_ins  ph   cert
    ("EAVE-CP-SE-001",  "eave",         0.08,   "EN_ISO_14683", "closed_panel", True,   False,  50,     False, "EN ISO 14683:2017 Table B.4"),
    ("EAVE-CP-SE-002",  "eave",         0.05,   "SVEBY",        "closed_panel", True,   True,   50,     False, "SVEBY Byggnadstypiska köldbryggor 2021"),
    ("CORNER-EXT-001",  "corner",       0.05,   "SVEBY",        "closed_panel", True,   False,  0,      False, "SVEBY 2021 Table 3.2"),
    ("CILL-WIN-SE-001", "cill",         0.04,   "SVEBY",        "closed_panel", True,   True,   0,      False, "SVEBY 2021"),
    ("JAMB-WIN-SE-001", "jamb",         0.02,   "SVEBY",        "closed_panel", True,   True,   0,      False, "SVEBY 2021"),
    ("HEAD-WIN-SE-001", "head",         0.01,   "SVEBY",        "closed_panel", True,   True,   0,      False, "SVEBY 2021"),
    ("FL-WALL-SE-001",  "floor_wall",   0.10,   "EN_ISO_14683", "closed_panel", False,  False,  0,      False, "EN ISO 14683:2017 Table B.3"),
    ("RIDGE-DUO-001",   "ridge",        0.04,   "SINTEF",       "closed_panel", True,   False,  0,      False, "SINTEF Byggforsk 471.015"),
    ("PARAPET-001",     "parapet",      0.15,   "EN_ISO_14683", "closed_panel", False,  False,  80,     False, "EN ISO 14683:2017"),
    ("FLOOR-EXT-001",   "floor_ground", 0.10,   "EN_ISO_14683", "closed_panel", False,  False,  0,      False, "EN ISO 14683:2017 Table B.5"),
]


def style_header_row(ws, col_defs):
    for col_idx, (name, required, width) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = BOLD_WHITE
        cell.fill = REQ_FILL if required else OPT_FILL
        cell.alignment = CENTRE
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30


def build_materials_sheet(wb):
    ws = wb.create_sheet("Materials")
    style_header_row(ws, MAT_COLS)
    col_names = [c[0] for c in MAT_COLS]
    for row_data in MATERIALS:
        row = {}
        row["name"]                          = row_data[0]
        row["manufacturer"]                  = row_data[1]
        row["spec_ref"]                      = row_data[2]
        row["lambda_W_mK"]                   = row_data[3]
        row["density_kg_m3"]                 = row_data[4]
        row["cp_J_kgK"]                      = row_data[5]
        row["fire_euroclass"]                = row_data[6]
        row["embodied_carbon_kgCO2e_per_kg"] = row_data[7]
        row["price_per_unit"]                = row_data[8]
        row["unit"]                          = row_data[9]
        row["currency"]                      = row_data[10]
        row["supplier_ref"]                  = row_data[11]
        ws.append([row.get(c) for c in col_names])


def build_junctions_sheet(wb):
    ws = wb.create_sheet("Junctions")
    style_header_row(ws, JNC_COLS)
    col_names = [c[0] for c in JNC_COLS]
    for jd in JUNCTIONS:
        row = dict(zip(col_names, jd))
        ws.append([row.get(c) for c in col_names])


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet
    build_materials_sheet(wb)
    build_junctions_sheet(wb)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")
    print(f"  Materials sheet : {len(MATERIALS)} rows")
    print(f"  Junctions sheet : {len(JUNCTIONS)} rows")
    print(f"\nEdit the file to add your own materials, then run:")
    print(f"  python tools/import_materials.py --file tools/materials_template.xlsx")


if __name__ == "__main__":
    main()
