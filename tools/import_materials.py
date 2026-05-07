"""Material library import tool.

Reads materials_template.xlsx and inserts rows into MaterialLibrary and
JunctionDetailLibrary under a given library_version_id.

Usage:
    python tools/import_materials.py --file tools/materials_template.xlsx --lib-version 1

Sheet: Materials
    Required columns: name, lambda_W_mK
    Optional: manufacturer, spec_ref, density_kg_m3, cp_J_kgK, fire_euroclass,
              embodied_carbon_kgCO2e_per_kg, price_per_unit, unit, currency, supplier_ref

Sheet: Junctions
    Required columns: code, type, psi_value_W_mK, psi_source
    Optional: build_up_type, insulation_continuity, thermal_break_present,
              min_outboard_insulation_mm, passivhaus_flag, cert_ref
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from dotenv import load_dotenv

load_dotenv()

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models import JunctionDetailLibrary, LibraryVersion, MaterialLibrary


def _str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ("TRUE", "YES", "1", "Y")
    return bool(val)


def import_materials(wb, db, lib_version_id: int) -> tuple[int, int, list[str]]:
    """Returns (inserted, skipped, skip_reasons)."""
    if "Materials" not in wb.sheetnames:
        return 0, 0, ["Sheet 'Materials' not found in workbook"]

    ws = wb["Materials"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col(row, name: str):
        try:
            idx = headers.index(name)
            return row[idx].value
        except ValueError:
            return None

    inserted = 0
    skipped = 0
    reasons: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = _str(col(row, "name"))
        lambda_val = _float(col(row, "lambda_W_mK"))

        if not name:
            skipped += 1
            reasons.append(f"Row {i}: missing 'name' — skipped")
            continue
        if lambda_val is None:
            skipped += 1
            reasons.append(f"Row {i} ({name}): missing or invalid 'lambda_W_mK' — skipped")
            continue

        mat = MaterialLibrary(
            library_version_id=lib_version_id,
            name=name,
            manufacturer=_str(col(row, "manufacturer")),
            spec_ref=_str(col(row, "spec_ref")),
            lambda_W_mK=lambda_val,
            density_kg_m3=_float(col(row, "density_kg_m3")),
            cp_J_kgK=_float(col(row, "cp_J_kgK")),
            fire_euroclass=_str(col(row, "fire_euroclass")),
            embodied_carbon_kgCO2e_per_kg=_float(col(row, "embodied_carbon_kgCO2e_per_kg")),
            price_per_unit=_float(col(row, "price_per_unit")),
            unit=_str(col(row, "unit")),
            currency=_str(col(row, "currency")) or "EUR",
            supplier_ref=_str(col(row, "supplier_ref")),
        )
        db.add(mat)
        inserted += 1

    db.flush()
    return inserted, skipped, reasons


def import_junctions(wb, db, lib_version_id: int) -> tuple[int, int, list[str]]:
    if "Junctions" not in wb.sheetnames:
        print("  No 'Junctions' sheet found — skipping junction import.")
        return 0, 0, []

    ws = wb["Junctions"]
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col(row, name: str):
        try:
            idx = headers.index(name)
            return row[idx].value
        except ValueError:
            return None

    inserted = 0
    skipped = 0
    reasons: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        code = _str(col(row, "code"))
        jtype = _str(col(row, "type"))
        psi = _float(col(row, "psi_value_W_mK"))
        source = _str(col(row, "psi_source"))

        if not code:
            skipped += 1
            reasons.append(f"Row {i}: missing 'code' — skipped")
            continue
        if psi is None:
            skipped += 1
            reasons.append(f"Row {i} ({code}): missing or invalid 'psi_value_W_mK' — skipped")
            continue
        if not source:
            skipped += 1
            reasons.append(f"Row {i} ({code}): missing 'psi_source' — skipped")
            continue

        junc = JunctionDetailLibrary(
            library_version_id=lib_version_id,
            code=code,
            type=jtype or "unknown",
            build_up_type=_str(col(row, "build_up_type")),
            insulation_continuity=_bool(col(row, "insulation_continuity") or True),
            thermal_break_present=_bool(col(row, "thermal_break_present") or False),
            min_outboard_insulation_mm=_float(col(row, "min_outboard_insulation_mm")),
            psi_value_W_mK=psi,
            psi_source=source,
            cert_ref=_str(col(row, "cert_ref")),
            passivhaus_flag=_bool(col(row, "passivhaus_flag") or False),
        )
        db.add(junc)
        inserted += 1

    db.flush()
    return inserted, skipped, reasons


def main():
    parser = argparse.ArgumentParser(description="Import materials from Excel into the database.")
    parser.add_argument("--file", required=True, help="Path to the Excel file")
    parser.add_argument("--lib-version", type=int, default=1, help="LibraryVersion.id to assign (default: 1)")
    parser.add_argument("--dry-run", action="store_true", help="Validate without inserting")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"Error: file not found: {args.file}")
        sys.exit(1)

    engine = create_engine(os.environ["DATABASE_URL"])
    Session = sessionmaker(bind=engine)

    with Session() as db:
        lib = db.get(LibraryVersion, args.lib_version)
        if not lib:
            print(f"Error: LibraryVersion id={args.lib_version} not found. Run seeds/library_v1.py first.")
            sys.exit(1)

        print(f"Importing into library '{lib.version}' (id={lib.id})")

        wb = openpyxl.load_workbook(args.file, data_only=True)

        mat_in, mat_sk, mat_reasons = import_materials(wb, db, lib.id)
        junc_in, junc_sk, junc_reasons = import_junctions(wb, db, lib.id)

        all_reasons = mat_reasons + junc_reasons
        if all_reasons:
            print("\nSkipped rows:")
            for r in all_reasons:
                print(f"  {r}")

        if args.dry_run:
            db.rollback()
            print(f"\nDry run — no changes written.")
        else:
            db.commit()

        print(f"\nMaterials:  {mat_in} inserted, {mat_sk} skipped")
        print(f"Junctions:  {junc_in} inserted, {junc_sk} skipped")


if __name__ == "__main__":
    main()
